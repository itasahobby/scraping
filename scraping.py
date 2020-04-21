#!/usr/bin/python3

import requests
import webbrowser
from bs4 import BeautifulSoup
import argparse
from tabulate import tabulate
from printy import printy
from printy import inputy
from openpyxl import Workbook
from openpyxl.styles import PatternFill,NamedStyle, Border, Side

def get_parser():
    parser = argparse.ArgumentParser(description="Scrape products from pccomponentes")
    parser.add_argument('-s',action="store", dest="search", help="Query to search from pccomponentes")
    parser.add_argument('-p',action="store_true", dest="print", help="Prints the results")
    parser.add_argument('-i',action="store_true", dest="interactive", help="Makes an interactive shell")
    parser.add_argument('-d',action="store", dest="dump", help="Dumps the result into an xml", metavar="FILENAME")    
    return parser

class Product:
    def __init__(self, name, url, brand, price):
        self.__name = name
        self.__url = url
        self.__brand = brand
        self.__price = price
    def toList(self):
        return [self.__name, self.__brand, self.__price]
    def toDict(self):
        return {
            "name": self.__name,
            "url": self.__url,
            "brand": self.__brand,
            "price": self.__price
        }
    def getName(self):
        return self.__name
    def getUrl(self):
        return self.__url
    def getBrand(self):
        return self.__brand
    def getPrice(self):
        return self.__price
    def setName(self,name):
        self.__name = name
    def setUrl(self,url):
        self.__url = url
    def setBrand(self,brand):
        self.__brand = brand
    def setPrice(self,price):
        self.__price = price

def scrape(target):
    html = requests.get(target)
    soup = BeautifulSoup(html.content,"html.parser")
    products = list(map(lambda product: product.find("a"), soup.find_all("article")))
    products_parsed = list(map(lambda product: Product(product["data-name"],target + product["href"],product["data-brand"],product["data-price"]),products))
    return products_parsed


def printTable(products):
    products_list = []
    i = 0
    for product in products:
        products_list.append([str(i)] + product.toList())
        i += 1
    printy(tabulate(products_list, headers=["","Name","brand","price"],tablefmt="fancy_grid"),"y")

def open_link(products):
    val = inputy("Enter a product number to open: ","b")
    val_int = int(val)
    if (val_int > 0) and (val_int < len(products)):
        printy(f"Opening link {products[val_int].getUrl()}","y")
        webbrowser.open(products[val_int].getUrl())
        return True
    else:
        return False

def printerr(url):
    printy(f"There is no results for {url}","r")

def interface(args,products):
    if(args.print):
        printTable(products)
    if(args.interactive and args.print):
        if not open_link(products):
            interface(args,products)

def get_header_style():
    style = NamedStyle("header")
    style.fill = PatternFill("solid", fgColor="7FCEEF")
    bd = Side(style="thick", color="000000")
    style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    return style

def get_product_style():
    style = NamedStyle("product")
    style.fill = PatternFill("solid", fgColor="BFE7F7")
    bd = Side(style="thin",color="000000")
    style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    return style

def dump(filename,products):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PcComponentes"
    first_product = products[0].toDict()
    header = [*first_product]
    column = 1
    header_style = get_header_style()
    for val in header:
        sheet.cell(1,column).value=val.capitalize()
        sheet.cell(1,column).style = header_style
        column += 1
    row = 2
    product_style = get_product_style()
    for product in products:
        sheet.cell(row,1).value=product.getName()
        sheet.cell(row,2).value=product.getUrl()
        sheet.cell(row,3).value=product.getBrand()
        sheet.cell(row,4).value=product.getPrice()
        for i in range(1,5):
            sheet.cell(row,i).style=product_style
        row += 1
    workbook.save(f"{filename}.xlsx")


def main():
    parser = get_parser()
    args = parser.parse_args()
    target = "https://www.pccomponentes.com" 
    if(args.search):
        target += f"/buscar/?query={args.search}" 
    products = scrape(target)
    if(len(products) > 0):
        interface(args,products)
        dump(args.dump,products)
    else:
        printerr(target)

main()